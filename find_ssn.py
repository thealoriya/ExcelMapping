from openpyxl import Workbook, load_workbook
import pandas as pd
wb = load_workbook("validation.xlsx")
ws = wb.active
# import numpy as np
# import glob
# files = glob.glob("*.xls*")
# print(files)
main = pd.DataFrame()
for r in range(2, 5):
    cntrl_id = ws.cell(r, 1).value
    df = pd.read_excel(cntrl_id)
    ssn = ws.cell(r, 2).value
    df = df[df.eq(ssn).any(1)]
    # if df.empty:
    #     print(False) # for a result List
    # else:
    #     print(True)
    main = main.append(df,ignore_index=True) # for DataFrame
    r = r+1;
print(main)